import os
import csv
import sqlite3
import anthropic
from dotenv import load_dotenv

load_dotenv()
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box

console = Console(force_terminal=True, color_system="truecolor")

# ─── TRANSLATIONS ──────────────────────────────────────────

TRANSLATIONS = {
    "darija": {
        # menu
        "menu_title":       "✨ HISSAB — Menu Principal",
        "option_0":         "🌐 Bdel langue",
        "option_1":         "💰 Zid revenu",
        "option_2":         "💸 Zid mssaref",
        "option_3":         "📊 Chouf rapport",
        "option_4":         "🤖 Tahlil Claude AI (mock)",
        "option_5":         "🤖 Tahlil Claude AI (real)",
        "option_6":         "💬 Chat m3a AI",
        "option_7":         "📁 Export CSV",
        "option_8":         "📁 Export Excel",
        "option_9":         "🚪 Khrej",
        "menu_prompt":      "Khtar option (0-9):",
        # inputs
        "montant_prompt":   "Montant (DH):",
        "desc_prompt":      "Description:",
        "question_prompt":  "Su'al dialk:",
        # confirmations
        "success_revenu":   "✅ Revenu {montant:.2f} DH tzad b najah!",
        "success_mssref":   "✅ Mssref {montant:.2f} DH tzad b najah!",
        "seed_data":        "✅ Data test tzadt automatiquement.",
        "langue_changed":   "🌐 Langue bdlat l: {label}",
        "csv_saved":        "📄 CSV tsajjlat f:\n{path}",
        "excel_saved":      "📊 Excel tsajjlat f:\n{path}",
        "exit_msg":         "بسلامة! 👋",
        "loading":          "Claude kaytlawn...",
        # errors
        "error_montant":    "❌ Dkhel raqm s7i7 — mithal: 1500 aw 1500.50",
        "error_option":     "❌ Option machi sahi9a — khtar bin 0 w 9.",
        "error_question":   "❌ Ma dkheltich hta su'al.",
        "error_credits":    "⚠️ Ma 3andkch credits — rouh console.anthropic.com tzid credits.",
        "error_apikey":     "⚠️ API key machi sahi9a — check ANTHROPIC_API_KEY dyalk.",
        "error_other":      "⚠️ Khata: {msg}",
        # panel titles
        "panel_mock":       "🤖 Tahlil Claude AI — Mock",
        "panel_real":       "🤖 Tahlil Claude AI — Real",
        "panel_chat":       "💬 Chat AI",
        # rapport
        "rapport_title":    "📊 Rapport Hissab",
        "col_indicateur":   "Indicateur",
        "col_montant":      "Montant (DH)",
        "row_revenus":      "💰 Total Revenus",
        "row_mssaref":      "💸 Total Mssaref",
        "row_rbah":         "Rbah / Khasara",
        # mock tahlil
        "mock_excellent":   "ممتاز ✅",
        "mock_bad":         "صعب ⚠️",
        "mock_text": (
            "[bold]1. 📊 تقييم الوضع المالي:[/bold]\n"
            "   الوضع المالي {statut}. هامش الربح {marge:.1f}%\n"
            "   مع إيرادات {rev:.2f} DH ومصاريف {ms:.2f} DH.\n\n"
            "[bold]2. 💪 أكبر نقطة قوة:[/bold]\n"
            "   تنوع مصادر الدخل يحمي المشروع من أي انهيار مفاجئ.\n\n"
            "[bold]3. ⚠️ أكبر نقطة ضعف:[/bold]\n"
            "   المصاريف الثابتة مرتفعة — أي انخفاض في الإيرادات سيؤثر مباشرة.\n\n"
            "[bold]4. 💡 توصية:[/bold]\n"
            "   زيد الإيرادات بـ 20% قبل ما تزيد أي مصاريف جديدة،\n"
            "   وتابع هامش الربح كل شهر باش تشوف التطور."
        ),
        "prompt_lang": "jaweb b darija maghribiya",
        # chat
        "chat_marge":         "📊 هامش الربح ديالك هو [bold]{marge:.1f}%[/] — {statut}",
        "chat_marge_ex":      "ممتاز 🟢",
        "chat_marge_ok":      "معقول 🟡",
        "chat_marge_bad":     "خطر 🔴",
        "chat_rbah_pos":      "💰 ربحتي [bold]{rbah:.2f} DH[/] — هامش الربح {marge:.1f}%. مزيان!",
        "chat_rbah_neg":      "⚠️ عندك خسارة [bold]{rbah:.2f} DH[/]. خاصك تراجع المصاريف.",
        "chat_mssaref":       "📋 المصاريف ديالك: {details}\nالمجموع: [bold]{total:.2f} DH[/]",
        "chat_mssaref_top":   "\n🔴 أكبر مصرف: [bold]{name}[/] ({val:.2f} DH)",
        "chat_revenus":       "📈 الإيرادات ديالك: {details}\nالمجموع: [bold]{total:.2f} DH[/]",
        "chat_nassiha_high":  "💡 وضعك مزيان — فكر في التوسع أو الاستثمار في منتج جديد.",
        "chat_nassiha_mid":   "💡 حاول تقلل المصاريف الثابتة بـ 10% وتزيد قناة دخل إضافية.",
        "chat_nassiha_low":   "💡 [bold]الأولوية:[/] قلل المصاريف فوراً وراجع كل مصرف واحد واحد.",
        "chat_wda3":          "📊 الوضع المالي: {statut}\nإيرادات: [bold]{rev:.2f} DH[/] | مصاريف: [bold]{ms:.2f} DH[/] | ربح: [bold]{rbah:.2f} DH[/]",
        "chat_wda3_ex":       "ممتاز 🟢",
        "chat_wda3_ok":       "معقول 🟡",
        "chat_wda3_bad":      "صعب 🔴",
        "chat_unknown":       "🤖 ما فهمتكش مزيان. سولني على: ربح، مصاريف، دخل، هامش، نصيحة، أو وضع.",
    },
    "english": {
        "menu_title":       "✨ HISSAB — Main Menu",
        "option_0":         "🌐 Change language",
        "option_1":         "💰 Add revenue",
        "option_2":         "💸 Add expense",
        "option_3":         "📊 View report",
        "option_4":         "🤖 AI Analysis (mock)",
        "option_5":         "🤖 AI Analysis (real)",
        "option_6":         "💬 Chat with AI",
        "option_7":         "📁 Export CSV",
        "option_8":         "📁 Export Excel",
        "option_9":         "🚪 Exit",
        "menu_prompt":      "Choose option (0-9):",
        "montant_prompt":   "Amount (DH):",
        "desc_prompt":      "Description:",
        "question_prompt":  "Your question:",
        "success_revenu":   "✅ Revenue of {montant:.2f} DH added successfully!",
        "success_mssref":   "✅ Expense of {montant:.2f} DH added successfully!",
        "seed_data":        "✅ Test data loaded automatically.",
        "langue_changed":   "🌐 Language changed to: {label}",
        "csv_saved":        "📄 CSV saved to:\n{path}",
        "excel_saved":      "📊 Excel saved to:\n{path}",
        "exit_msg":         "Goodbye! 👋",
        "loading":          "Claude is thinking...",
        "error_montant":    "❌ Please enter a valid number — e.g. 1500 or 1500.50",
        "error_option":     "❌ Invalid option — choose between 0 and 9.",
        "error_question":   "❌ You didn't enter a question.",
        "error_credits":    "⚠️ No credits — go to console.anthropic.com to add credits.",
        "error_apikey":     "⚠️ Invalid API key — check your ANTHROPIC_API_KEY.",
        "error_other":      "⚠️ Error: {msg}",
        "panel_mock":       "🤖 Claude AI Analysis — Mock",
        "panel_real":       "🤖 Claude AI Analysis — Real",
        "panel_chat":       "💬 Chat AI",
        "rapport_title":    "📊 Financial Report",
        "col_indicateur":   "Indicator",
        "col_montant":      "Amount (DH)",
        "row_revenus":      "💰 Total Revenue",
        "row_mssaref":      "💸 Total Expenses",
        "row_rbah":         "Profit / Loss",
        "mock_excellent":   "Excellent ✅",
        "mock_bad":         "Difficult ⚠️",
        "mock_text": (
            "[bold]1. 📊 Financial Overview:[/bold]\n"
            "   Status: {statut}. Profit margin: {marge:.1f}%\n"
            "   Revenue: {rev:.2f} DH | Expenses: {ms:.2f} DH.\n\n"
            "[bold]2. 💪 Key Strength:[/bold]\n"
            "   Diversified income protects the business from sudden downturns.\n\n"
            "[bold]3. ⚠️ Key Weakness:[/bold]\n"
            "   Fixed costs are high — any drop in revenue impacts margins directly.\n\n"
            "[bold]4. 💡 Recommendation:[/bold]\n"
            "   Grow revenue by 20% before adding new expenses,\n"
            "   and track profit margin monthly to measure progress."
        ),
        "prompt_lang":        "Respond in English",
        "chat_marge":         "📊 Your profit margin is [bold]{marge:.1f}%[/] — {statut}",
        "chat_marge_ex":      "Excellent 🟢",
        "chat_marge_ok":      "Fair 🟡",
        "chat_marge_bad":     "Danger 🔴",
        "chat_rbah_pos":      "💰 You made [bold]{rbah:.2f} DH[/] in profit — margin {marge:.1f}%. Great!",
        "chat_rbah_neg":      "⚠️ You have a loss of [bold]{rbah:.2f} DH[/]. Review your expenses.",
        "chat_mssaref":       "📋 Your expenses: {details}\nTotal: [bold]{total:.2f} DH[/]",
        "chat_mssaref_top":   "\n🔴 Biggest expense: [bold]{name}[/] ({val:.2f} DH)",
        "chat_revenus":       "📈 Your revenue: {details}\nTotal: [bold]{total:.2f} DH[/]",
        "chat_nassiha_high":  "💡 Finances look great — consider expanding or investing in a new product.",
        "chat_nassiha_mid":   "💡 Try reducing fixed costs by 10% and add an additional income stream.",
        "chat_nassiha_low":   "💡 [bold]Priority:[/] Cut expenses immediately and review each one carefully.",
        "chat_wda3":          "📊 Financial status: {statut}\nRevenue: [bold]{rev:.2f} DH[/] | Expenses: [bold]{ms:.2f} DH[/] | Profit: [bold]{rbah:.2f} DH[/]",
        "chat_wda3_ex":       "Excellent 🟢",
        "chat_wda3_ok":       "Fair 🟡",
        "chat_wda3_bad":      "Poor 🔴",
        "chat_unknown":       "🤖 I didn't understand. Ask me about: profit, expenses, revenue, margin, advice, or status.",
    },
    "arabic": {
        "menu_title":       "✨ حساب — القائمة الرئيسية",
        "option_0":         "🌐 تغيير اللغة",
        "option_1":         "💰 إضافة دخل",
        "option_2":         "💸 إضافة مصروف",
        "option_3":         "📊 عرض التقرير",
        "option_4":         "🤖 تحليل الذكاء الاصطناعي (تجريبي)",
        "option_5":         "🤖 تحليل الذكاء الاصطناعي (حقيقي)",
        "option_6":         "💬 المحادثة مع الذكاء الاصطناعي",
        "option_7":         "📁 تصدير CSV",
        "option_8":         "📁 تصدير Excel",
        "option_9":         "🚪 خروج",
        "menu_prompt":      "اختر خياراً (0-9):",
        "montant_prompt":   "المبلغ (درهم):",
        "desc_prompt":      "الوصف:",
        "question_prompt":  "سؤالك:",
        "success_revenu":   "✅ تمت إضافة دخل بقيمة {montant:.2f} درهم بنجاح!",
        "success_mssref":   "✅ تمت إضافة مصروف بقيمة {montant:.2f} درهم بنجاح!",
        "seed_data":        "✅ تم تحميل البيانات التجريبية تلقائياً.",
        "langue_changed":   "🌐 تم تغيير اللغة إلى: {label}",
        "csv_saved":        "📄 تم حفظ ملف CSV في:\n{path}",
        "excel_saved":      "📊 تم حفظ ملف Excel في:\n{path}",
        "exit_msg":         "مع السلامة! 👋",
        "loading":          "الذكاء الاصطناعي يفكر...",
        "error_montant":    "❌ أدخل رقماً صحيحاً — مثال: 1500 أو 1500.50",
        "error_option":     "❌ خيار غير صحيح — اختر بين 0 و 9.",
        "error_question":   "❌ لم تدخل أي سؤال.",
        "error_credits":    "⚠️ رصيدك منتهٍ — اذهب إلى console.anthropic.com لإضافة رصيد.",
        "error_apikey":     "⚠️ مفتاح API غير صحيح — تحقق من ANTHROPIC_API_KEY.",
        "error_other":      "⚠️ خطأ: {msg}",
        "panel_mock":       "🤖 تحليل كلود — تجريبي",
        "panel_real":       "🤖 تحليل كلود — حقيقي",
        "panel_chat":       "💬 المحادثة مع الذكاء الاصطناعي",
        "rapport_title":    "📊 التقرير المالي",
        "col_indicateur":   "المؤشر",
        "col_montant":      "المبلغ (درهم)",
        "row_revenus":      "💰 إجمالي الإيرادات",
        "row_mssaref":      "💸 إجمالي المصاريف",
        "row_rbah":         "الربح / الخسارة",
        "mock_excellent":   "ممتاز ✅",
        "mock_bad":         "صعب ⚠️",
        "mock_text": (
            "[bold]1. 📊 تقييم الوضع المالي:[/bold]\n"
            "   الوضع المالي {statut}. هامش الربح {marge:.1f}%\n"
            "   الإيرادات: {rev:.2f} درهم | المصاريف: {ms:.2f} درهم.\n\n"
            "[bold]2. 💪 أبرز نقطة قوة:[/bold]\n"
            "   تنوع مصادر الدخل يحمي المشروع من أي انهيار مفاجئ.\n\n"
            "[bold]3. ⚠️ أبرز نقطة ضعف:[/bold]\n"
            "   المصاريف الثابتة مرتفعة — أي انخفاض في الإيرادات سيؤثر مباشرةً.\n\n"
            "[bold]4. 💡 توصية:[/bold]\n"
            "   زِد إيراداتك بنسبة 20% قبل إضافة أي مصاريف جديدة،\n"
            "   وتابع هامش الربح شهرياً لرصد التطور."
        ),
        "prompt_lang":        "أجب بالعربية الفصحى",
        "chat_marge":         "📊 هامش الربح لديك هو [bold]{marge:.1f}%[/] — {statut}",
        "chat_marge_ex":      "ممتاز 🟢",
        "chat_marge_ok":      "مقبول 🟡",
        "chat_marge_bad":     "خطر 🔴",
        "chat_rbah_pos":      "💰 حققت ربحاً قدره [bold]{rbah:.2f} درهم[/] — هامش الربح {marge:.1f}%. ممتاز!",
        "chat_rbah_neg":      "⚠️ لديك خسارة بقيمة [bold]{rbah:.2f} درهم[/]. يجب مراجعة المصاريف.",
        "chat_mssaref":       "📋 مصاريفك: {details}\nالمجموع: [bold]{total:.2f} درهم[/]",
        "chat_mssaref_top":   "\n🔴 أكبر مصروف: [bold]{name}[/] ({val:.2f} درهم)",
        "chat_revenus":       "📈 إيراداتك: {details}\nالمجموع: [bold]{total:.2f} درهم[/]",
        "chat_nassiha_high":  "💡 وضعك المالي ممتاز — فكّر في التوسع أو الاستثمار في منتج جديد.",
        "chat_nassiha_mid":   "💡 حاول تخفيض التكاليف الثابتة بنسبة 10% وأضف مصدر دخل إضافياً.",
        "chat_nassiha_low":   "💡 [bold]الأولوية:[/] خفّض المصاريف فوراً وراجع كل بند على حدة.",
        "chat_wda3":          "📊 الوضع المالي: {statut}\nالإيرادات: [bold]{rev:.2f} درهم[/] | المصاريف: [bold]{ms:.2f} درهم[/] | الربح: [bold]{rbah:.2f} درهم[/]",
        "chat_wda3_ex":       "ممتاز 🟢",
        "chat_wda3_ok":       "مقبول 🟡",
        "chat_wda3_bad":      "ضعيف 🔴",
        "chat_unknown":       "🤖 لم أفهم سؤالك. اسألني عن: الربح، المصاريف، الإيرادات، هامش الربح، نصيحة، أو الوضع العام.",
    },
}

LANGUE_LABELS = {"darija": "🇲🇦 Darija", "english": "🇬🇧 English", "arabic": "🇸🇦 عربي"}
LANGUE_CYCLE  = ["darija", "english", "arabic"]


class Hissab:
    def __init__(self, db_path: str = "hissab.db"):
        self._db_path = db_path
        self._client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        self.langue = "darija"
        self._init_db()

    def t(self, key: str, **kwargs) -> str:
        text = TRANSLATIONS[self.langue][key]
        return text.format(**kwargs) if kwargs else text

    # ─── DB ───────────────────────────────────────────────

    def _conn(self):
        return sqlite3.connect(self._db_path)

    def _init_db(self):
        with self._conn() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS revenus (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    montant     REAL    NOT NULL,
                    description TEXT    DEFAULT '',
                    date        TEXT    DEFAULT (datetime('now'))
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS mssaref (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    montant     REAL    NOT NULL,
                    description TEXT    DEFAULT '',
                    date        TEXT    DEFAULT (datetime('now'))
                )
            """)

    # ─── INSERT ───────────────────────────────────────────

    def zid_revenu(self, montant: float, description: str = ""):
        with self._conn() as conn:
            conn.execute(
                "INSERT INTO revenus (montant, description) VALUES (?, ?)",
                (montant, description),
            )

    def zid_mssref(self, montant: float, description: str = ""):
        with self._conn() as conn:
            conn.execute(
                "INSERT INTO mssaref (montant, description) VALUES (?, ?)",
                (montant, description),
            )

    # ─── SELECT ───────────────────────────────────────────

    def get_all_revenus(self) -> list:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT id, montant, description, date FROM revenus ORDER BY date"
            ).fetchall()
        return [{"id": r[0], "montant": r[1], "description": r[2], "date": r[3]} for r in rows]

    def get_all_mssaref(self) -> list:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT id, montant, description, date FROM mssaref ORDER BY date"
            ).fetchall()
        return [{"id": r[0], "montant": r[1], "description": r[2], "date": r[3]} for r in rows]

    def total_revenus(self) -> float:
        with self._conn() as conn:
            result = conn.execute("SELECT COALESCE(SUM(montant), 0) FROM revenus").fetchone()
        return result[0]

    def total_mssaref(self) -> float:
        with self._conn() as conn:
            result = conn.execute("SELECT COALESCE(SUM(montant), 0) FROM mssaref").fetchone()
        return result[0]

    def rbah(self) -> float:
        return self.total_revenus() - self.total_mssaref()

    # ─── RAPPORT ──────────────────────────────────────────

    def rapport(self):
        rbah = self.rbah()
        rbah_icon = "✅" if rbah >= 0 else "⚠️"

        table = Table(
            title=self.t("rapport_title"),
            box=box.ROUNDED,
            title_style="bold",
            header_style="bold",
            min_width=42,
        )
        table.add_column(self.t("col_indicateur"), style="bold", width=22)
        table.add_column(self.t("col_montant"), justify="right", width=18)
        table.add_row(self.t("row_revenus"), f"{self.total_revenus():,.2f}")
        table.add_row(self.t("row_mssaref"), f"{self.total_mssaref():,.2f}")
        table.add_section()
        table.add_row(f"{rbah_icon} {self.t('row_rbah')}", f"[bold]{rbah:,.2f}[/]")
        console.print()
        console.print(table)

    # ─── CLAUDE AI ────────────────────────────────────────

    def tahlil_bi_claude(self, mock: bool = False) -> str:
        if mock:
            return self._mock_tahlil()

        revenus_details = "\n".join(
            f"  - {r['description']}: {r['montant']:.2f} DH" for r in self.get_all_revenus()
        )
        mssaref_details = "\n".join(
            f"  - {m['description']}: {m['montant']:.2f} DH" for m in self.get_all_mssaref()
        )

        prompt = (
            f"You are an expert financial advisor. Analyze this data and give practical recommendations. "
            f"{self.t('prompt_lang')}.\n\n"
            f"Revenue:\n{revenus_details}\nTotal: {self.total_revenus():.2f} DH\n\n"
            f"Expenses:\n{mssaref_details}\nTotal: {self.total_mssaref():.2f} DH\n\n"
            f"Net Profit: {self.rbah():.2f} DH\n\n"
            f"Provide: 1. Quick assessment  2. Biggest strength  3. Biggest weakness  4. One recommendation"
        )

        message = self._client.messages.create(
            model="claude-opus-4-8",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        return message.content[0].text

    def _mock_tahlil(self) -> str:
        rbah  = self.rbah()
        marge = (rbah / self.total_revenus() * 100) if self.total_revenus() else 0
        statut = self.t("mock_excellent") if rbah > 0 else self.t("mock_bad")
        return self.t("mock_text").format(
            statut=statut, marge=marge,
            rev=self.total_revenus(), ms=self.total_mssaref()
        )

    # ─── CHAT ─────────────────────────────────────────────

    def chat_with_ai(self, question: str) -> str:
        q     = question.lower().strip()
        rbah  = self.rbah()
        marge = (rbah / self.total_revenus() * 100) if self.total_revenus() else 0

        if any(w in q for w in ["marge", "هامش", "margin", "نسبة"]):
            if marge > 50:   statut = self.t("chat_marge_ex")
            elif marge > 20: statut = self.t("chat_marge_ok")
            else:            statut = self.t("chat_marge_bad")
            return self.t("chat_marge", marge=marge, statut=statut)

        if any(w in q for w in ["rbah", "ربح", "ribh", "profit", "benefice"]):
            if rbah > 0:
                return self.t("chat_rbah_pos", rbah=rbah, marge=marge)
            else:
                return self.t("chat_rbah_neg", rbah=abs(rbah))

        if any(w in q for w in ["mssaref", "مصاريف", "مصرف", "expense", "depense"]):
            mssaref = self.get_all_mssaref()
            top     = max(mssaref, key=lambda x: x["montant"], default=None)
            details = ", ".join(f"{m['description']} ({m['montant']:.0f})" for m in mssaref)
            rep     = self.t("chat_mssaref", details=details, total=self.total_mssaref())
            if top:
                rep += self.t("chat_mssaref_top", name=top["description"], val=top["montant"])
            return rep

        if any(w in q for w in ["revenus", "دخل", "إيرادات", "revenue", "income", "dakhl"]):
            revenus = self.get_all_revenus()
            details = ", ".join(f"{r['description']} ({r['montant']:.0f})" for r in revenus)
            return self.t("chat_revenus", details=details, total=self.total_revenus())

        if any(w in q for w in ["nassiha", "نصيحة", "advice", "conseil", "توصية"]):
            if marge > 60:   return self.t("chat_nassiha_high")
            elif marge > 30: return self.t("chat_nassiha_mid")
            else:            return self.t("chat_nassiha_low")

        if any(w in q for w in ["wda3", "وضع", "كيفاش", "status", "situation"]):
            if marge > 60:   statut = self.t("chat_wda3_ex")
            elif marge > 20: statut = self.t("chat_wda3_ok")
            else:            statut = self.t("chat_wda3_bad")
            return self.t("chat_wda3", statut=statut,
                          rev=self.total_revenus(), ms=self.total_mssaref(), rbah=rbah)

        return self.t("chat_unknown")

    # ─── EXPORT ───────────────────────────────────────────

    def export_csv(self, filename: str = "hissab_export.csv") -> str:
        path = os.path.abspath(filename)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["type", "montant", "description", "date"])
            for r in self.get_all_revenus():
                writer.writerow(["revenu", r["montant"], r["description"], r["date"]])
            for m in self.get_all_mssaref():
                writer.writerow(["mssref", m["montant"], m["description"], m["date"]])
            writer.writerow([])
            writer.writerow(["", "TOTAL REVENUS", "", self.total_revenus()])
            writer.writerow(["", "TOTAL MSSAREF", "", self.total_mssaref()])
            writer.writerow(["", "RBAH", "", self.rbah()])
        return path

    def export_excel(self, filename: str = "hissab_export.xlsx") -> str:
        wb = Workbook()
        hf = Font(bold=True, color="FFFFFF")

        def _header(ws, headers, color):
            ws.append(headers)
            fill = PatternFill("solid", fgColor=color)
            for cell in ws[1]:
                cell.font = hf
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

        def _widths(ws, widths):
            for col, w in zip("ABCD", widths):
                ws.column_dimensions[col].width = w

        ws_r = wb.active
        ws_r.title = "Revenus"
        _header(ws_r, ["ID", "Montant (DH)", "Description", "Date"], "1F7A4D")
        for r in self.get_all_revenus():
            ws_r.append([r["id"], r["montant"], r["description"], r["date"]])
        _widths(ws_r, [6, 15, 30, 22])

        ws_m = wb.create_sheet("Mssaref")
        _header(ws_m, ["ID", "Montant (DH)", "Description", "Date"], "C0392B")
        for m in self.get_all_mssaref():
            ws_m.append([m["id"], m["montant"], m["description"], m["date"]])
        _widths(ws_m, [6, 15, 30, 22])

        ws_s = wb.create_sheet("Resume")
        _header(ws_s, ["Indicateur", "Montant (DH)"], "2471A3")
        rbah = self.rbah()
        ws_s.append(["Total Revenus", self.total_revenus()])
        ws_s.append(["Total Mssaref", self.total_mssaref()])
        ws_s.append(["Rbah / Khasara", rbah])
        ws_s["B4"].font = Font(bold=True, color="1F7A4D" if rbah >= 0 else "C0392B")
        ws_s.column_dimensions["A"].width = 20
        ws_s.column_dimensions["B"].width = 16

        path = os.path.abspath(filename)
        wb.save(path)
        return path


# ─── MENU ─────────────────────────────────────────────────

def _ask_montant(h: Hissab) -> float:
    while True:
        try:
            return float(console.input(f"   {h.t('montant_prompt')} ").strip())
        except ValueError:
            console.print(Panel(h.t("error_montant")))


def _print_menu(h: Hissab):
    label = LANGUE_LABELS[h.langue]
    lines = (
        f"  [bold]0[/].  {h.t('option_0')}  [dim]({label})[/dim]\n"
        f"  [bold]1[/].  {h.t('option_1')}\n"
        f"  [bold]2[/].  {h.t('option_2')}\n"
        f"  [bold]3[/].  {h.t('option_3')}\n"
        f"  [bold]4[/].  {h.t('option_4')}\n"
        f"  [bold]5[/].  {h.t('option_5')}\n"
        f"  [bold]6[/].  {h.t('option_6')}\n"
        f"  [bold]7[/].  {h.t('option_7')}\n"
        f"  [bold]8[/].  {h.t('option_8')}\n"
        f"  [bold]9[/].  {h.t('option_9')}"
    )
    console.print(Panel(lines, title=f"[bold]{h.t('menu_title')}[/]", padding=(1, 4)))


def main_menu():
    h = Hissab()

    if h.total_revenus() == 0 and h.total_mssaref() == 0:
        h.zid_revenu(15000, "Ventes janvier")
        h.zid_revenu(8000, "Prestation service")
        h.zid_mssref(3000, "Loyer")
        h.zid_mssref(1200, "Salaires")
        h.zid_mssref(500, "Electricite")
        console.print(Panel(h.t("seed_data")))

    while True:
        console.print()
        _print_menu(h)
        choix = console.input(f"\n  [bold]{h.t('menu_prompt')}[/] ").strip()

        if choix == "0":
            idx = LANGUE_CYCLE.index(h.langue)
            h.langue = LANGUE_CYCLE[(idx + 1) % len(LANGUE_CYCLE)]
            console.print(Panel(h.t("langue_changed", label=LANGUE_LABELS[h.langue])))

        elif choix == "1":
            console.print()
            montant = _ask_montant(h)
            desc = console.input(f"   {h.t('desc_prompt')} ").strip()
            h.zid_revenu(montant, desc)
            console.print(Panel(h.t("success_revenu", montant=montant)))

        elif choix == "2":
            console.print()
            montant = _ask_montant(h)
            desc = console.input(f"   {h.t('desc_prompt')} ").strip()
            h.zid_mssref(montant, desc)
            console.print(Panel(h.t("success_mssref", montant=montant)))

        elif choix == "3":
            h.rapport()

        elif choix == "4":
            result = h.tahlil_bi_claude(mock=True)
            console.print(Panel(result, title=f"[bold]{h.t('panel_mock')}[/]", padding=(1, 2)))

        elif choix == "5":
            try:
                with console.status(f"[bold]{h.t('loading')}[/]", spinner="dots"):
                    result = h.tahlil_bi_claude(mock=False)
                console.print(Panel(result, title=f"[bold]{h.t('panel_real')}[/]", padding=(1, 2)))
            except Exception as e:
                msg = str(e).lower()
                if "credit" in msg or "balance" in msg:
                    console.print(Panel(h.t("error_credits")))
                elif "api" in msg or "auth" in msg or "key" in msg:
                    console.print(Panel(h.t("error_apikey")))
                else:
                    console.print(Panel(h.t("error_other", msg=e)))

        elif choix == "6":
            console.print()
            su2al = console.input(f"   [bold]{h.t('question_prompt')}[/] ").strip()
            if su2al:
                result = h.chat_with_ai(su2al)
                console.print(Panel(result, title=f"[bold]{h.t('panel_chat')}[/]", padding=(1, 2)))
            else:
                console.print(Panel(h.t("error_question")))

        elif choix == "7":
            path = h.export_csv()
            console.print(Panel(h.t("csv_saved", path=path)))

        elif choix == "8":
            path = h.export_excel()
            console.print(Panel(h.t("excel_saved", path=path)))

        elif choix == "9":
            console.print(Panel(f"[bold]{h.t('exit_msg')}[/]"))
            break

        else:
            console.print(Panel(h.t("error_option")))


if __name__ == "__main__":
    try:
        main_menu()
    except KeyboardInterrupt:
        h = Hissab()
        console.print(Panel(f"[bold]\n{h.t('exit_msg')}[/]"))
